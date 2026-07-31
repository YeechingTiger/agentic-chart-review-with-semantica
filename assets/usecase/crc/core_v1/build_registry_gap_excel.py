#!/usr/bin/env python3
"""Build a reviewable CRC guideline-rule/variable/registry Excel workbook."""

from __future__ import annotations

import argparse
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


BLOCKS = ("eligibility", "action", "timing", "exceptions")
AVAILABILITY = {
    "registry_direct": "Yes — direct registry field",
    "registry_coarsened": "Partial — registry scope mismatch",
    "chart_extension": "No — linked-chart field needed",
    "outside_current_sources": "No — new source needed",
    "derived": "Derived — not stored directly",
}
STATUS_FILL = {
    "Yes": "C6EFCE",
    "Partial": "FFEB9C",
    "No": "FFC7CE",
    "Derived": "DDEBF7",
    "NOT_ASSESSED": "E7E6E6",
}
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
WHITE_BOLD = Font(color="FFFFFF", bold=True)
THIN_GRAY = Side(style="thin", color="D9E1F2")
CELL_BORDER = Border(bottom=THIN_GRAY)


def load_yaml(path: Path) -> dict[str, Any]:
    value = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    if not isinstance(value, dict):
        raise ValueError(f"{path}: expected a mapping")
    return value


def as_list(value: Any) -> list[Any]:
    return value if isinstance(value, list) else []


def cell_text(value: Any) -> str:
    """Render structured YAML values as readable, deterministic Excel text."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (str, int, float)):
        return str(value)
    if isinstance(value, list):
        return "\n".join(cell_text(item) for item in value)
    if isinstance(value, dict):
        return "\n".join(
            f"{key}: {cell_text(item)}" for key, item in value.items()
        )
    return str(value)


def ast_has_unresolved(node: Any) -> bool:
    if not isinstance(node, dict):
        return False
    if node.get("op") == "unresolved":
        return True
    for operand in as_list(node.get("operands")):
        if ast_has_unresolved(operand):
            return True
    return any(
        ast_has_unresolved(node.get(key))
        for key in ("operand", "condition", "then", "else")
    )


def registry_item(variable: dict[str, Any]) -> tuple[str, str]:
    mapping = variable.get("registry_mapping") or {}
    if not isinstance(mapping, dict):
        return "", ""
    name = str(mapping.get("item_name") or "")
    number = str(mapping.get("item_number") or "")
    if name == "none":
        name = ""
    if number == "none":
        number = ""
    return name, number


def scope_mismatch(
    variable_id: str, projection_by_variable: dict[str, dict[str, Any]]
) -> str:
    projection = projection_by_variable.get(variable_id) or {}
    return "\n".join(
        str(value) for value in as_list(projection.get("explicit_loss_list"))
    )


def rule_check_text(rule: dict[str, Any]) -> str:
    requirements = rule["requirements"]
    return "\n".join(
        [
            f"Eligibility: {requirements['eligibility']['expression']}",
            f"Action: {requirements['action']['expression']}",
            f"Timing: {requirements['timing']['expression']}",
            f"Exceptions: {requirements['exceptions']['expression']}",
        ]
    )


def style_table_sheet(
    worksheet: Any,
    *,
    table_name: str,
    freeze: str,
    widths: dict[str, float],
    availability_headers: tuple[str, ...] = (),
) -> None:
    worksheet.freeze_panes = freeze
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_view.showGridLines = False
    worksheet.row_dimensions[1].height = 32
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = CELL_BORDER
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width
    if worksheet.max_row >= 2:
        table = Table(displayName=table_name, ref=worksheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

    header_by_name = {
        str(cell.value): cell.column_letter for cell in worksheet[1] if cell.value
    }
    for header in availability_headers:
        column = header_by_name.get(header)
        if not column:
            continue
        cell_range = f"{column}2:{column}{worksheet.max_row}"
        for prefix, color in STATUS_FILL.items():
            worksheet.conditional_formatting.add(
                cell_range,
                FormulaRule(
                    formula=[f'LEFT({column}2,{len(prefix)})="{prefix}"'],
                    fill=PatternFill("solid", fgColor=color),
                ),
            )


def build_workbook(bundle: Path, output: Path) -> dict[str, int]:
    normalization = bundle / "normalization"
    rules_doc = load_yaml(normalization / "normalized_rules.yaml")
    variables_doc = load_yaml(normalization / "canonical_variables.yaml")
    projections_doc = load_yaml(normalization / "registry_projections.yaml")
    coverage_doc = load_yaml(normalization / "evidence_coverage.yaml")

    rules = as_list(rules_doc.get("rules"))
    variables = as_list(variables_doc.get("variables"))
    variable_by_id = {str(row["variable_id"]): row for row in variables}
    projection_by_variable = {
        str(row["canonical_variable_id"]): row
        for row in as_list(projections_doc.get("projections"))
    }
    structural_rule_by_id = {
        str(row["candidate_id"]): row for row in as_list(coverage_doc.get("rules"))
    }

    workbook = Workbook()
    workbook.remove(workbook.active)

    rules_sheet = workbook.create_sheet("Rules")
    rule_headers = [
        "Category",
        "Check Rule ID",
        "Check Rule",
        "Variables Needed",
        "Available in Cancer Registry?",
        "Direct Variables",
        "Scope-Mismatch Variables",
        "Missing Linked-Chart Variables",
        "Missing New-Source Variables",
        "Derived Variables",
        "Registry Component Summary",
        "Computability",
        "Blockers",
        "Source ID",
        "Source Anchor",
        "Current CRC Data Observed?",
    ]
    rules_sheet.append(rule_headers)

    detail_sheet = workbook.create_sheet("Rule_Variable_Detail")
    detail_headers = [
        "Category",
        "Check Rule ID",
        "Check Rule Title",
        "Requirement Role",
        "Check Clause",
        "Predicate Logic",
        "Variable Needed",
        "Variable Label",
        "Available in Cancer Registry?",
        "Mapping Level",
        "Registry Item",
        "Registry Item Number",
        "Temporal Meaning",
        "Missingness Semantics",
        "Scope Mismatch / Information Loss",
        "Rule Computability",
        "Rule Blockers",
        "Variable Contract",
        "Current CRC Data Observed?",
    ]
    detail_sheet.append(detail_headers)

    variable_usage: dict[str, list[dict[str, str]]] = defaultdict(list)
    rule_mapping_counts: Counter[str] = Counter()
    variable_use_count = 0

    for rule in sorted(
        rules, key=lambda row: (str(row["category"]), str(row["candidate_id"]))
    ):
        unique_variable_ids: list[str] = []
        seen_variables: set[str] = set()
        mapping_groups: dict[str, list[str]] = defaultdict(list)
        has_unresolved = False
        for block in BLOCKS:
            requirement = rule["requirements"][block]
            has_unresolved = has_unresolved or ast_has_unresolved(
                requirement.get("evidence_logic")
            )
            for variable_id_raw in as_list(requirement.get("variables")):
                variable_id = str(variable_id_raw)
                variable = variable_by_id[variable_id]
                mapping_level = str(variable["mapping_level"])
                if variable_id not in seen_variables:
                    unique_variable_ids.append(variable_id)
                    seen_variables.add(variable_id)
                    mapping_groups[mapping_level].append(variable_id)

                item_name, item_number = registry_item(variable)
                computability = str(
                    (rule.get("computability") or {}).get("status") or ""
                )
                blockers = "\n".join(
                    str(value)
                    for value in as_list(
                        (rule.get("computability") or {}).get("blockers")
                    )
                )
                contract_name = (
                    "CRC.FULL.VAR."
                    + variable_id.removeprefix("crc.").replace(".", "__")
                    + ".yaml"
                )
                detail_sheet.append(
                    [
                        rule["category"],
                        rule["candidate_id"],
                        rule["title"],
                        block,
                        requirement["expression"],
                        (rule.get("source_context") or {}).get("predicate_logic", ""),
                        variable_id,
                        variable.get("label", ""),
                        AVAILABILITY[mapping_level],
                        mapping_level,
                        item_name,
                        item_number,
                        cell_text(variable.get("temporal_meaning")),
                        cell_text(variable.get("missingness_semantics")),
                        scope_mismatch(variable_id, projection_by_variable),
                        computability,
                        blockers,
                        str(normalization / "variable_contracts" / contract_name),
                        "NOT_ASSESSED",
                    ]
                )
                variable_usage[variable_id].append(
                    {
                        "candidate_id": str(rule["candidate_id"]),
                        "category": str(rule["category"]),
                        "role": block,
                    }
                )
                variable_use_count += 1

        structural = structural_rule_by_id.get(str(rule["candidate_id"])) or {}
        strict_state = str(
            structural.get("nonconcordance_defensibility") or "none"
        )
        if strict_state == "full":
            whole_rule = "Yes — fully available"
        elif strict_state == "partial":
            whole_rule = "Partial — not sufficient for full rule"
        else:
            whole_rule = "No — full rule not available"
        if has_unresolved:
            whole_rule += "; unresolved rule block"
        rule_mapping_counts[whole_rule.split(" — ", 1)[0]] += 1

        variable_lines = [
            f"{variable_id} — {AVAILABILITY[str(variable_by_id[variable_id]['mapping_level'])]}"
            for variable_id in unique_variable_ids
        ]
        direct = mapping_groups["registry_direct"]
        coarsened = mapping_groups["registry_coarsened"]
        chart = mapping_groups["chart_extension"]
        outside = mapping_groups["outside_current_sources"]
        derived = mapping_groups["derived"]
        component_summary = (
            f"direct={len(direct)}; scope_mismatch={len(coarsened)}; "
            f"linked_chart_missing={len(chart)}; new_source_missing={len(outside)}; "
            f"derived={len(derived)}"
        )
        rules_sheet.append(
            [
                rule["category"],
                rule["candidate_id"],
                f"{rule['title']}\n\n{rule_check_text(rule)}\n\n"
                f"Predicate: {(rule.get('source_context') or {}).get('predicate_logic', '')}",
                "\n".join(variable_lines),
                whole_rule,
                "\n".join(direct),
                "\n".join(coarsened),
                "\n".join(chart),
                "\n".join(outside),
                "\n".join(derived),
                component_summary,
                (rule.get("computability") or {}).get("status", ""),
                "\n".join(
                    str(value)
                    for value in as_list(
                        (rule.get("computability") or {}).get("blockers")
                    )
                ),
                rule.get("source_id", ""),
                rule.get("source_anchor", ""),
                "NOT_ASSESSED",
            ]
        )

    variable_sheet = workbook.create_sheet("Variable_Summary")
    variable_headers = [
        "Variable",
        "Label",
        "Rule-Block Uses",
        "Distinct Rules",
        "Categories",
        "Requirement Roles",
        "Available in Cancer Registry?",
        "Mapping Level",
        "Registry Item",
        "Registry Item Number",
        "Temporal Meaning",
        "Missingness Semantics",
        "Scope Mismatch / Information Loss",
        "Example Rules",
        "Current CRC Data Observed?",
    ]
    variable_sheet.append(variable_headers)
    for variable in sorted(
        variables,
        key=lambda row: (
            AVAILABILITY[str(row["mapping_level"])],
            -len(variable_usage[str(row["variable_id"])]),
            str(row["variable_id"]),
        ),
    ):
        variable_id = str(variable["variable_id"])
        uses = variable_usage[variable_id]
        candidate_ids = sorted({use["candidate_id"] for use in uses})
        item_name, item_number = registry_item(variable)
        variable_sheet.append(
            [
                variable_id,
                variable.get("label", ""),
                len(uses),
                len(candidate_ids),
                ", ".join(sorted({use["category"] for use in uses})),
                ", ".join(sorted({use["role"] for use in uses})),
                AVAILABILITY[str(variable["mapping_level"])],
                variable["mapping_level"],
                item_name,
                item_number,
                cell_text(variable.get("temporal_meaning")),
                cell_text(variable.get("missingness_semantics")),
                scope_mismatch(variable_id, projection_by_variable),
                ", ".join(candidate_ids[:12])
                + (f", +{len(candidate_ids) - 12} more" if len(candidate_ids) > 12 else ""),
                "NOT_ASSESSED",
            ]
        )

    category_sheet = workbook.create_sheet("Category_Summary")
    category_headers = [
        "Category",
        "Rules",
        "Distinct Variables",
        "Registry Direct",
        "Registry Scope Mismatch",
        "Missing Linked-Chart",
        "Missing New Source",
        "Derived",
        "Whole Rules Available",
    ]
    category_sheet.append(category_headers)
    category_rules: dict[str, set[str]] = defaultdict(set)
    category_variables: dict[str, set[str]] = defaultdict(set)
    for variable_id, uses in variable_usage.items():
        for use in uses:
            category_rules[use["category"]].add(use["candidate_id"])
            category_variables[use["category"]].add(variable_id)
    for category in sorted(category_rules):
        ids = category_variables[category]
        counts = Counter(
            str(variable_by_id[variable_id]["mapping_level"]) for variable_id in ids
        )
        category_sheet.append(
            [
                category,
                len(category_rules[category]),
                len(ids),
                counts["registry_direct"],
                counts["registry_coarsened"],
                counts["chart_extension"],
                counts["outside_current_sources"],
                counts["derived"],
                0,
            ]
        )

    legend_sheet = workbook.create_sheet("Read_Me", 0)
    legend_rows = [
        ["CRC guideline rule-variable-registry matrix", ""],
        ["Clinical use", "NOT FOR CLINICAL USE — candidate authoring assessment"],
        ["Rule denominator", len(rules)],
        ["Canonical variables", len(variables)],
        ["Rule-variable rows", variable_use_count],
        ["Observed CRC data", "NOT_ASSESSED"],
        [
            "Availability basis",
            "Candidate NAACCR/STORE structural mapping, not measured availability in a CRC tumor-level extract.",
        ],
        ["", ""],
        ["Status", "Meaning"],
        [
            AVAILABILITY["registry_direct"],
            "Candidate exact canonical-to-registry mapping; still requires registrar and clinical review.",
        ],
        [
            AVAILABILITY["registry_coarsened"],
            "A registry item exists but loses clinical, temporal, provenance, or value-domain detail.",
        ],
        [
            AVAILABILITY["chart_extension"],
            "No accepted registry field; extract from linked chart documents/events.",
        ],
        [
            AVAILABILITY["outside_current_sources"],
            "Needs a policy, label, calendar, institutional metric, or other new source.",
        ],
        [
            AVAILABILITY["derived"],
            "Not stored directly; compute from reviewed inputs with provenance.",
        ],
        ["NOT_ASSESSED", "No compatible tumor-level CRC linked-registry profile is currently bound."],
        ["", ""],
        ["Sheet", "Purpose"],
        ["Rules", "One row per guideline candidate; the first five columns match the requested view."],
        ["Rule_Variable_Detail", "One row per rule requirement role and variable (2,406 rows)."],
        ["Variable_Summary", "One row per canonical variable (576 rows)."],
        ["Category_Summary", "Counts by guideline category."],
    ]
    for row in legend_rows:
        legend_sheet.append(row)
    legend_sheet.sheet_view.showGridLines = False
    legend_sheet.column_dimensions["A"].width = 36
    legend_sheet.column_dimensions["B"].width = 115
    legend_sheet.freeze_panes = "A9"
    for cell in legend_sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
    for cell in legend_sheet[9]:
        cell.fill = SUBHEADER_FILL
        cell.font = Font(bold=True)
    for cell in legend_sheet[18]:
        cell.fill = SUBHEADER_FILL
        cell.font = Font(bold=True)
    for row in legend_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row_index in range(10, 16):
        value = str(legend_sheet.cell(row_index, 1).value)
        for prefix, color in STATUS_FILL.items():
            if value.startswith(prefix):
                legend_sheet.cell(row_index, 1).fill = PatternFill(
                    "solid", fgColor=color
                )

    style_table_sheet(
        rules_sheet,
        table_name="RuleRegistryMatrix",
        freeze="A2",
        widths={
            "A": 28,
            "B": 22,
            "C": 70,
            "D": 68,
            "E": 32,
            "F": 30,
            "G": 34,
            "H": 42,
            "I": 36,
            "J": 30,
            "K": 42,
            "L": 20,
            "M": 42,
            "N": 30,
            "O": 38,
            "P": 24,
        },
        availability_headers=(
            "Available in Cancer Registry?",
            "Current CRC Data Observed?",
        ),
    )
    style_table_sheet(
        detail_sheet,
        table_name="RuleVariableDetail",
        freeze="A2",
        widths={
            "A": 25,
            "B": 22,
            "C": 48,
            "D": 18,
            "E": 48,
            "F": 48,
            "G": 45,
            "H": 36,
            "I": 32,
            "J": 25,
            "K": 34,
            "L": 18,
            "M": 34,
            "N": 48,
            "O": 58,
            "P": 22,
            "Q": 42,
            "R": 58,
            "S": 24,
        },
        availability_headers=(
            "Available in Cancer Registry?",
            "Current CRC Data Observed?",
        ),
    )
    style_table_sheet(
        variable_sheet,
        table_name="VariableRegistrySummary",
        freeze="A2",
        widths={
            "A": 48,
            "B": 38,
            "C": 18,
            "D": 16,
            "E": 52,
            "F": 24,
            "G": 32,
            "H": 25,
            "I": 35,
            "J": 18,
            "K": 38,
            "L": 52,
            "M": 60,
            "N": 58,
            "O": 24,
        },
        availability_headers=(
            "Available in Cancer Registry?",
            "Current CRC Data Observed?",
        ),
    )
    style_table_sheet(
        category_sheet,
        table_name="CategoryRegistrySummary",
        freeze="A2",
        widths={
            "A": 30,
            "B": 14,
            "C": 20,
            "D": 18,
            "E": 26,
            "F": 25,
            "G": 23,
            "H": 14,
            "I": 24,
        },
    )

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)

    # Reopen the file to ensure the produced OOXML archive is readable.
    verified = load_workbook(output, read_only=True, data_only=False)
    expected_sheets = {
        "Read_Me",
        "Rules",
        "Rule_Variable_Detail",
        "Variable_Summary",
        "Category_Summary",
    }
    if set(verified.sheetnames) != expected_sheets:
        raise ValueError(f"unexpected workbook sheets: {verified.sheetnames}")
    if verified["Rules"].max_row != len(rules) + 1:
        raise ValueError("Rules sheet row count mismatch")
    if verified["Rule_Variable_Detail"].max_row != variable_use_count + 1:
        raise ValueError("Rule_Variable_Detail row count mismatch")
    if verified["Variable_Summary"].max_row != len(variables) + 1:
        raise ValueError("Variable_Summary row count mismatch")
    verified.close()

    return {
        "rules": len(rules),
        "variables": len(variables),
        "variable_uses": variable_use_count,
        "categories": len(category_rules),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "bundle",
        type=Path,
        help="Path to authoring/crc/core_v1",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Output .xlsx path; defaults inside the bundle.",
    )
    args = parser.parse_args()
    bundle = args.bundle.resolve()
    output = (
        args.output.resolve()
        if args.output
        else bundle / "CRC_guideline_rule_variable_registry_matrix.xlsx"
    )
    summary = build_workbook(bundle, output)
    print(
        f"wrote {output}: {summary['rules']} rules, {summary['variables']} variables, "
        f"{summary['variable_uses']} rule-variable rows, {summary['categories']} categories"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
